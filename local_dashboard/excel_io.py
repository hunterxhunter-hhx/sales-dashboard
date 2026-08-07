from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook


def iter_xlsx_rows(path: Path) -> Iterator[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        # Some exported workbooks contain a stale <dimension> of A1 even though
        # the sheet XML has all rows. In read_only mode openpyxl trusts that
        # cache unless dimensions are reset.
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
        for values in rows:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[index] if index < len(values) else ""
            if any(value not in (None, "") for value in row.values()):
                yield row
    finally:
        workbook.close()
