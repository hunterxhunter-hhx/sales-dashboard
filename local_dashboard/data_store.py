from __future__ import annotations

import json
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path
from typing import Any


OWNER_MAP = {
    "589": "祁春如",
    "586": "赵宇欣",
    "314": "辛雨薇",
    "56": "黄依婷",
    "260": "万雪莲",
    "585": "吴秋禅",
    "135": "徐佳莹",
    "80": "石玉",
    "633": "魏富欣",
    "61": "徐佳莹",
}

CHANNEL_RULES = [
    ("1-挂图承接【私域部】", "1-挂图承接【私域部】"),
    ("2-公号-引流", "2-公号-引流"),
    ("2-企微-回捞", "2-企微-回捞"),
    ("2-企微-询课", "2-企微-询课"),
    ("2-有赞-下单", "2-有赞-下单"),
    ("2-视频号-下单", "2-视频号-下单"),
]

MISSING_CHANNELS = {"", "缺失渠道", "待归因", "待补录"}

SOURCE_FILE_SPECS = [
    ("carry", "用户承接表.xlsx"),
    ("orders", "用户订单表.xlsx"),
]

FORECAST_DETAIL_FILENAME = "明细表.xlsx"
CHANNEL_CLASSIFICATION_VERSION = 2
YOUZAN_FILENAME = "有赞数据.xlsx"


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def pick(row: dict[str, Any], names: list[str]) -> Any:
    lower_map = {safe_text(key).lower().replace(" ", "").replace("_", ""): key for key in row}
    for name in names:
        if name in row:
            return row[name]
        compact = name.lower().replace(" ", "").replace("_", "")
        if compact in lower_map:
            return row[lower_map[compact]]
    return ""


def parse_datetime(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = safe_text(value)
    if not text:
        return ""
    text = text.replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text


def to_float(value: Any) -> float:
    text = safe_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def excel_serial_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
    parsed = parse_datetime(value)
    return parsed[:10] if parsed else ""


def period_date(value: Any) -> tuple[str, bool]:
    text = safe_text(value)
    match = re.match(r"^(\d{2})(\d{2})(\d{2})$", text)
    if not match:
        return "", False
    year, month, day = match.groups()
    return f"20{year}-{month}-{day}", True


def owner_name(employee_user_id: str, owner_raw: str) -> str:
    if employee_user_id in OWNER_MAP:
        return OWNER_MAP[employee_user_id]
    for key, value in OWNER_MAP.items():
        if re.search(rf"(^|[^0-9]){re.escape(key)}([^0-9]|$)", owner_raw):
            return value
    return owner_raw or "缺失企微"


def classify_channel(tags: Any, remark: Any = "") -> str:
    remark_text = safe_text(remark)
    if remark_text:
        return remark_text

    text = safe_text(tags)
    hits = [label for needle, label in CHANNEL_RULES if needle in text]
    if not hits:
        return "一转"
    if len(hits) > 1:
        return "冲突渠道"
    return hits[0]


def classify_scenario(product_name: Any) -> str:
    return "直播间转化" if "直播" in safe_text(product_name) else "追单转化"


def attribution_channel(channel: Any, customer_tags: Any) -> str:
    value = safe_text(channel)
    if value and value not in MISSING_CHANNELS:
        return value
    return classify_channel(customer_tags)


def apply_manual_remark_channel(channel: Any, manual_remark: Any) -> str:
    value = safe_text(channel)
    if value not in MISSING_CHANNELS:
        return value
    return safe_text(manual_remark) or value


def is_activation_promoter(value: Any) -> bool:
    return safe_text(value) == "常丁健"


def month_sort_key(value: Any) -> tuple[int, int, str]:
    text_value = safe_text(value)
    match = re.search(r"(\d{4})[^\d]?(\d{1,2})", text_value)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month, text_value
    date_text = parse_datetime(value)[:7]
    if re.match(r"^\d{4}-\d{2}$", date_text):
        year, month = date_text.split("-")
        return int(year), int(month), text_value
    return 0, 0, text_value


def date_sort_key(value: Any) -> tuple[str, str]:
    text_value = safe_text(value)
    return text_value[:10], text_value


def normalize_carry_row(row: dict[str, Any]) -> dict[str, Any]:
    union_id = safe_text(pick(row, ["Union ID", "unionid", "union_id", "UnionID"]))
    employee_user_id = safe_text(
        pick(row, ["员工 user_id", "员工user_id", "员工 user id", "员工ID", "员工微信id", "所属员工ID"])
    )
    owner_raw = safe_text(pick(row, ["所属员工", "员工", "所属销售", "员工微信名称"]))
    if not employee_user_id:
        prefix = owner_raw.split("-", 1)[0].strip()
        if prefix.isdigit():
            employee_user_id = prefix
    add_time = parse_datetime(pick(row, ["添加时间", "首次添加时间", "创建时间"]))
    customer_tags = safe_text(pick(row, ["客户标签", "标签"]))
    carry_remark = safe_text(pick(row, ["备注", "备注列", "客户备注", "渠道备注"]))
    return {
        "union_id": union_id,
        "employee_user_id": employee_user_id,
        "add_time": add_time,
        "owner_raw": owner_raw,
        "owner_name": owner_name(employee_user_id, owner_raw),
        "channel": classify_channel(customer_tags, carry_remark),
        "customer_tags": customer_tags,
        "raw_json": json.dumps(row, ensure_ascii=False, default=str),
    }


def normalize_order_row(row: dict[str, Any]) -> dict[str, Any]:
    order_id = safe_text(pick(row, ["订单ID", "订单id", "订单 Id", "订单号"]))
    order_no = safe_text(pick(row, ["订单号", "订单编号"]))
    wxid = safe_text(pick(row, ["微信id", "微信ID", "微信 id", "unionid"]))
    user_id = safe_text(pick(row, ["用户ID", "客户ID", "用户id", "客户id"]))
    order_time = parse_datetime(pick(row, ["支付时间", "下单时间", "付款时间"]))
    paid_amount = to_float(pick(row, ["实付金额", "支付金额", "订单实付金额"]))
    refund_amount = to_float(pick(row, ["退款金额", "售后退款金额"]))
    product_name = safe_text(pick(row, ["商品名称", "课程名称"]))
    promoter = safe_text(pick(row, ["推广员"]))
    promoter_modified = safe_text(pick(row, ["推广员（修改后）", "推广员(修改后)", "推广员修改后"]))
    manual_remark = safe_text(pick(row, ["手动备注"]))
    return {
        "order_id": order_id or order_no,
        "order_no": order_no,
        "wxid": wxid,
        "user_id": user_id,
        "order_time": order_time,
        "paid_amount": paid_amount,
        "refund_amount": refund_amount,
        "net_sales": max(paid_amount - refund_amount, 0.0),
        "order_status": safe_text(pick(row, ["订单状态"])),
        "product_name": product_name,
        "scenario": classify_scenario(product_name),
        "promoter": promoter,
        "promoter_modified": promoter_modified,
        "manual_remark": manual_remark,
        "raw_json": json.dumps(row, ensure_ascii=False, default=str),
    }


class DashboardStore:
    def __init__(self, db_path: Path, root_dir: Path | None = None) -> None:
        self.db_path = Path(db_path)
        self.root_dir = Path(root_dir) if root_dir else self.db_path.parent.parent

    def connect(self) -> sqlite3.Connection:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    @contextmanager
    def session(self):
        conn = self.connect()
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def init_db(self) -> None:
        self.root_dir.mkdir(parents=True, exist_ok=True)
        (self.root_dir / "data").mkdir(exist_ok=True)
        (self.root_dir / "data" / "source").mkdir(parents=True, exist_ok=True)
        (self.root_dir / "imports" / "carry").mkdir(parents=True, exist_ok=True)
        (self.root_dir / "imports" / "orders").mkdir(parents=True, exist_ok=True)
        (self.root_dir / "exports").mkdir(exist_ok=True)
        with self.session() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS import_batches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    data_type TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    rows_seen INTEGER NOT NULL DEFAULT 0,
                    inserted INTEGER NOT NULL DEFAULT 0,
                    updated INTEGER NOT NULL DEFAULT 0,
                    duplicates INTEGER NOT NULL DEFAULT 0,
                    errors INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS carry_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    union_id TEXT NOT NULL,
                    employee_user_id TEXT NOT NULL,
                    add_time TEXT NOT NULL,
                    owner_raw TEXT,
                    owner_name TEXT,
                    channel TEXT,
                    customer_tags TEXT,
                    raw_json TEXT,
                    batch_id INTEGER,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(union_id, employee_user_id, add_time)
                );

                CREATE TABLE IF NOT EXISTS order_records (
                    order_id TEXT PRIMARY KEY,
                    order_no TEXT,
                    wxid TEXT,
                    user_id TEXT,
                    order_time TEXT,
                    paid_amount REAL NOT NULL DEFAULT 0,
                    refund_amount REAL NOT NULL DEFAULT 0,
                    net_sales REAL NOT NULL DEFAULT 0,
                    order_status TEXT,
                    product_name TEXT,
                    scenario TEXT,
                    promoter TEXT,
                    promoter_modified TEXT,
                    manual_remark TEXT,
                    raw_json TEXT,
                    batch_id INTEGER,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS order_attribution_overrides (
                    order_id TEXT PRIMARY KEY,
                    employee_user_id TEXT NOT NULL,
                    owner_name TEXT NOT NULL,
                    reason TEXT,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS order_attributions (
                    order_id TEXT PRIMARY KEY,
                    union_id TEXT,
                    carry_record_id INTEGER,
                    employee_user_id TEXT,
                    owner_name TEXT,
                    channel TEXT,
                    attribution_type TEXT NOT NULL,
                    reason TEXT,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS source_files (
                    data_type TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    path TEXT NOT NULL,
                    size INTEGER NOT NULL,
                    mtime_ns INTEGER NOT NULL,
                    last_imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    last_result_json TEXT
                );
                """
            )
            self._ensure_column(conn, "order_records", "promoter", "TEXT")
            self._ensure_column(conn, "order_records", "promoter_modified", "TEXT")
            self._ensure_column(conn, "order_records", "manual_remark", "TEXT")

    def _ensure_column(self, conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
        columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
        if column not in columns:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    def source_dir(self) -> Path:
        return self.root_dir / "data" / "source"

    def _source_path(self, filename: str, source_dir: Path | None = None) -> Path:
        return (source_dir or self.source_dir()) / filename

    def forecast_detail_rows(self, source_dir: Path | None = None) -> list[dict[str, Any]]:
        from .excel_io import iter_xlsx_rows

        path = self._source_path(FORECAST_DETAIL_FILENAME, Path(source_dir) if source_dir else None)
        if not path.exists():
            return []
        rows = []
        for row in iter_xlsx_rows(path):
            period = safe_text(pick(row, ["期数", "日期", "时间"]))
            month_date = excel_serial_date(pick(row, ["月份", "月"]))
            date, is_daily = period_date(period)
            channel = safe_text(pick(row, ["渠道", "来源渠道"])) or "缺失渠道"
            owner = safe_text(pick(row, ["销售", "企微", "所属员工"])) or "待补录"
            if not month_date or not channel or not owner:
                continue
            rows.append(
                {
                    "period": period,
                    "date": date or month_date,
                    "month": month_date[:7],
                    "channel": channel,
                    "owner": owner,
                    "carry": to_float(pick(row, ["承接人数", "承接", "应销转"])),
                    "gmv": to_float(pick(row, ["gmv", "GMV", "销售额"])),
                    "attendance": to_float(pick(row, ["到课", "到课人数"])),
                    "is_daily": is_daily,
                }
            )
        return rows

    def youzan_summary(self, source_dir: Path | None = None) -> dict[str, Any]:
        from .excel_io import iter_xlsx_rows

        source_dir = Path(source_dir) if source_dir else self.source_dir()
        path = None
        for candidate in source_dir.iterdir():
            if candidate.is_file() and candidate.suffix.lower() == ".xlsx" and "有赞" in candidate.name:
                path = candidate
                break
        if path is None or not path.exists():
            return {"available": False, "sheets": {}, "cards": {}, "charts": {}}

        def as_number(value: Any) -> float:
            if isinstance(value, (int, float)):
                return float(value)
            text_value = safe_text(value).replace(",", "")
            if not text_value:
                return 0.0
            match = re.search(r"-?\d+(?:\.\d+)?", text_value)
            if not match:
                return 0.0
            try:
                return float(match.group(0))
            except ValueError:
                return 0.0

        def safe_month(value: Any) -> str:
            text = safe_text(value)
            if not text:
                return ""
            if re.match(r"^\d{4}-\d{1,2}$", text):
                year, month = text.split("-")
                return f"{int(year):04d}-{int(month):02d}"
            if re.match(r"^\d{4}/\d{1,2}$", text):
                year, month = text.split("/")
                return f"{int(year):04d}-{int(month):02d}"
            parsed = parse_datetime(text)
            return parsed[:7] if re.match(r"^\d{4}-\d{2}", parsed) else text

        workbook_rows = {}
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    if hasattr(sheet, "reset_dimensions"):
                        sheet.reset_dimensions()
                    rows = sheet.iter_rows(values_only=True)
                    headers = [safe_text(value) for value in next(rows, [])]
                    sheet_rows = []
                    for values in rows:
                        row = {}
                        for index, header in enumerate(headers):
                            if not header:
                                continue
                            row[header] = values[index] if index < len(values) else ""
                        if any(value not in (None, "") for value in row.values()):
                            sheet_rows.append(row)
                    workbook_rows[sheet_name] = sheet_rows
            finally:
                workbook.close()
        except Exception:
            workbook_rows = {sheet_name: list(iter_xlsx_rows(path, sheet_name)) for sheet_name in ["CRM导出订单-6", "订单表-6", "月毛利-6"]}

        crm_rows = workbook_rows.get("CRM导出订单-6", [])
        order_rows = workbook_rows.get("订单表-6", [])
        month_rows = workbook_rows.get("月毛利-6", [])

        month_entries = []
        for row in month_rows:
            month_value = safe_month(pick(row, ["月份"]))
            if not month_value:
                continue
            month_entries.append(
                {
                    "month": month_value,
                    "gmv": as_number(pick(row, ["GMV（除退款）"])),
                    "profit": as_number(pick(row, ["本月预估毛利（物流成本需月底核算）"])),
                    "orders": as_number(pick(row, ["订单数"])),
                }
            )

        latest_month_entry = None
        if month_entries:
            latest_month_entry = sorted(month_entries, key=lambda item: month_sort_key(item["month"]))[-1]

        month_latest = latest_month_entry or {"month": "", "gmv": 0.0, "profit": 0.0, "orders": 0.0}
        fallback_gmv = sum(as_number(pick(row, ["GMV（已减退款）"])) for row in order_rows)
        if not month_latest["gmv"]:
            month_latest["gmv"] = fallback_gmv

        sku_gmv = defaultdict(float)
        sku_profit = defaultdict(float)
        type_gmv = defaultdict(float)
        type_profit = defaultdict(float)
        daily = defaultdict(lambda: {"gmv": 0.0, "profit": 0.0, "orders": 0})

        for row in crm_rows:
            sku = safe_text(pick(row, ["商品简称"])) or "未命名SKU"
            gmv = as_number(pick(row, ["最终GMV（除退款）"]))
            profit = as_number(pick(row, ["订单利润"]))
            product_type = safe_text(pick(row, ["商品类型"])) or "未分类"
            sku_gmv[sku] += gmv
            sku_profit[sku] += profit
            type_gmv[product_type] += gmv
            type_profit[product_type] += profit

        for row in order_rows:
            day = safe_text(pick(row, ["付款日期"]))
            if not day:
                day = parse_datetime(pick(row, ["付款时间"]))[:10]
            if not day:
                continue
            bucket = daily[day]
            bucket["gmv"] += as_number(pick(row, ["GMV（已减退款）"]))
            bucket["profit"] += as_number(pick(row, ["预估利润（物流成本月末核算）"]))
            bucket["orders"] += 1

        def top_n_map(source: dict[str, float], limit: int = 10) -> list[dict[str, Any]]:
            items = sorted(source.items(), key=lambda item: (-item[1], item[0]))
            return [{"name": name, "value": round(value, 2)} for name, value in items[:limit]]

        daily_rows = [
            {"date": date, "gmv": values["gmv"], "profit": values["profit"], "orders": values["orders"]}
            for date, values in sorted(daily.items(), key=lambda item: date_sort_key(item[0]))
        ]

        return {
            "available": True,
            "sheets": {
                "crm": {"sheet": "CRM导出订单-6", "rows": len(crm_rows)},
                "orders": {"sheet": "订单表-6", "rows": len(order_rows)},
                "month": {"sheet": "月毛利-6", "rows": len(month_rows)},
            },
            "cards": {
                "estimated_month_gmv": round(month_latest["gmv"], 2),
                "estimated_month_profit": round(month_latest["profit"], 2),
            },
            "charts": {
                "sku_gmv": top_n_map(sku_gmv, 10),
                "sku_profit": top_n_map(sku_profit, 10),
                "type_gmv": [{"name": name, "value": round(value, 2)} for name, value in sorted(type_gmv.items(), key=lambda item: (-item[1], item[0]))],
                "type_profit": [{"name": name, "value": round(value, 2)} for name, value in sorted(type_profit.items(), key=lambda item: (-item[1], item[0]))],
                "daily": daily_rows,
            },
        }

    def _tracked_source(self, conn: sqlite3.Connection, data_type: str) -> sqlite3.Row | None:
        return conn.execute(
            "SELECT * FROM source_files WHERE data_type = ?",
            (data_type,),
        ).fetchone()

    def _order_source_is_consistent(self, conn: sqlite3.Connection, tracked: sqlite3.Row | None) -> bool:
        if not tracked:
            return False
        last_result = json.loads(tracked["last_result_json"] or "{}")
        source_rows_seen = last_result.get("rows_seen")
        if source_rows_seen is None:
            return True
        order_count = conn.execute("SELECT COUNT(*) AS count FROM order_records").fetchone()["count"]
        return int(source_rows_seen) == int(order_count)

    def source_file_status(self, source_dir: Path | None = None) -> dict[str, Any]:
        source_dir = Path(source_dir) if source_dir else self.source_dir()
        files = []
        with self.session() as conn:
            for data_type, filename in SOURCE_FILE_SPECS:
                path = self._source_path(filename, source_dir)
                tracked = self._tracked_source(conn, data_type)
                exists = path.exists()
                size = path.stat().st_size if exists else 0
                mtime_ns = path.stat().st_mtime_ns if exists else 0
                tracked_size = int(tracked["size"]) if tracked else 0
                tracked_mtime_ns = int(tracked["mtime_ns"]) if tracked else 0
                files.append(
                    {
                        "data_type": data_type,
                        "filename": filename,
                        "path": str(path),
                        "exists": exists,
                        "size": size,
                        "mtime_ns": mtime_ns,
                        "changed": exists and (not tracked or size != tracked_size or mtime_ns != tracked_mtime_ns),
                        "last_imported_at": tracked["last_imported_at"] if tracked else "",
                        "last_result": json.loads(tracked["last_result_json"] or "{}") if tracked else {},
                    }
                )
        return {
            "ready": all(item["exists"] for item in files),
            "source_dir": str(source_dir),
            "files": files,
        }

    def sync_source_files(self, source_dir: Path | None = None) -> dict[str, Any]:
        from .excel_io import iter_xlsx_rows

        source_dir = Path(source_dir) if source_dir else self.source_dir()
        source_dir.mkdir(parents=True, exist_ok=True)
        imports = []
        for data_type, filename in SOURCE_FILE_SPECS:
            path = self._source_path(filename, source_dir)
            if not path.exists():
                continue
            stat = path.stat()
            with self.session() as conn:
                tracked = self._tracked_source(conn, data_type)
                unchanged = tracked and int(tracked["size"]) == stat.st_size and int(tracked["mtime_ns"]) == stat.st_mtime_ns
                if data_type == "carry" and unchanged:
                    last_result = json.loads(tracked["last_result_json"] or "{}")
                    unchanged = int(last_result.get("channel_classification_version") or 0) == CHANNEL_CLASSIFICATION_VERSION
                if data_type == "orders" and unchanged:
                    unchanged = self._order_source_is_consistent(conn, tracked)
            if unchanged:
                continue

            if data_type == "carry":
                result = self.import_carry_rows(iter_xlsx_rows(path), filename)
            else:
                result = self.replace_order_rows(iter_xlsx_rows(path), filename)

            with self.session() as conn:
                conn.execute(
                    """
                    INSERT INTO source_files(data_type, filename, path, size, mtime_ns, last_result_json)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(data_type) DO UPDATE SET
                        filename = excluded.filename,
                        path = excluded.path,
                        size = excluded.size,
                        mtime_ns = excluded.mtime_ns,
                        last_imported_at = CURRENT_TIMESTAMP,
                        last_result_json = excluded.last_result_json
                    """,
                    (data_type, filename, str(path), stat.st_size, stat.st_mtime_ns, json.dumps(result, ensure_ascii=False)),
                )
            imports.append({"data_type": data_type, "filename": filename, "result": result})

        return {
            "changed": bool(imports),
            "imports": imports,
            "status": self.source_file_status(source_dir),
        }

    def _create_batch(self, conn: sqlite3.Connection, data_type: str, filename: str) -> int:
        cursor = conn.execute(
            "INSERT INTO import_batches(data_type, filename) VALUES (?, ?)",
            (data_type, filename),
        )
        return int(cursor.lastrowid)

    def _finish_batch(self, conn: sqlite3.Connection, batch_id: int, result: dict[str, int]) -> None:
        conn.execute(
            """
            UPDATE import_batches
            SET rows_seen = ?, inserted = ?, updated = ?, duplicates = ?, errors = ?
            WHERE id = ?
            """,
            (
                result["rows_seen"],
                result["inserted"],
                result["updated"],
                result["duplicates"],
                result["errors"],
                batch_id,
            ),
        )

    def import_carry_rows(self, rows, filename: str) -> dict[str, int]:
        result = {"rows_seen": 0, "inserted": 0, "updated": 0, "duplicates": 0, "errors": 0}
        with self.session() as conn:
            batch_id = self._create_batch(conn, "carry", filename)
            for row in rows:
                result["rows_seen"] += 1
                item = normalize_carry_row(row)
                if not item["union_id"] or not item["employee_user_id"] or not item["add_time"]:
                    result["errors"] += 1
                    continue
                existing = conn.execute(
                    """
                    SELECT owner_raw, owner_name, channel, customer_tags, raw_json
                    FROM carry_records
                    WHERE union_id = ? AND employee_user_id = ? AND add_time = ?
                    """,
                    (
                        item["union_id"],
                        item["employee_user_id"],
                        item["add_time"],
                    ),
                ).fetchone()
                if existing:
                    changed = any(
                        safe_text(existing[column]) != safe_text(item[column])
                        for column in ["owner_raw", "owner_name", "channel", "customer_tags", "raw_json"]
                    )
                    if changed:
                        conn.execute(
                            """
                            UPDATE carry_records
                            SET owner_raw = ?, owner_name = ?, channel = ?, customer_tags = ?,
                                raw_json = ?, batch_id = ?
                            WHERE union_id = ? AND employee_user_id = ? AND add_time = ?
                            """,
                            (
                                item["owner_raw"],
                                item["owner_name"],
                                item["channel"],
                                item["customer_tags"],
                                item["raw_json"],
                                batch_id,
                                item["union_id"],
                                item["employee_user_id"],
                                item["add_time"],
                            ),
                        )
                        result["updated"] += 1
                    else:
                        result["duplicates"] += 1
                else:
                    conn.execute(
                        """
                        INSERT INTO carry_records(
                            union_id, employee_user_id, add_time, owner_raw, owner_name,
                            channel, customer_tags, raw_json, batch_id
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item["union_id"],
                            item["employee_user_id"],
                            item["add_time"],
                            item["owner_raw"],
                            item["owner_name"],
                            item["channel"],
                            item["customer_tags"],
                            item["raw_json"],
                            batch_id,
                        ),
                    )
                    result["inserted"] += 1
            self._finish_batch(conn, batch_id, result)
        result["channel_classification_version"] = CHANNEL_CLASSIFICATION_VERSION
        self.recompute_attributions()
        return result

    def import_order_rows(self, rows, filename: str) -> dict[str, int]:
        result = {"rows_seen": 0, "inserted": 0, "updated": 0, "duplicates": 0, "errors": 0}
        with self.session() as conn:
            batch_id = self._create_batch(conn, "orders", filename)
            for row in rows:
                result["rows_seen"] += 1
                item = normalize_order_row(row)
                if not item["order_id"]:
                    result["errors"] += 1
                    continue
                exists = conn.execute(
                    "SELECT 1 FROM order_records WHERE order_id = ?",
                    (item["order_id"],),
                ).fetchone()
                conn.execute(
                    """
                    INSERT INTO order_records(
                        order_id, order_no, wxid, user_id, order_time, paid_amount,
                        refund_amount, net_sales, order_status, product_name, scenario,
                        promoter, promoter_modified, manual_remark, raw_json, batch_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(order_id) DO UPDATE SET
                        order_no = excluded.order_no,
                        wxid = excluded.wxid,
                        user_id = excluded.user_id,
                        order_time = excluded.order_time,
                        paid_amount = excluded.paid_amount,
                        refund_amount = excluded.refund_amount,
                        net_sales = excluded.net_sales,
                        order_status = excluded.order_status,
                        product_name = excluded.product_name,
                        scenario = excluded.scenario,
                        promoter = excluded.promoter,
                        promoter_modified = excluded.promoter_modified,
                        manual_remark = excluded.manual_remark,
                        raw_json = excluded.raw_json,
                        batch_id = excluded.batch_id,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (
                        item["order_id"],
                        item["order_no"],
                        item["wxid"],
                        item["user_id"],
                        item["order_time"],
                        item["paid_amount"],
                        item["refund_amount"],
                        item["net_sales"],
                        item["order_status"],
                        item["product_name"],
                        item["scenario"],
                        item["promoter"],
                        item["promoter_modified"],
                        item["manual_remark"],
                        item["raw_json"],
                        batch_id,
                    ),
                )
                if exists:
                    result["updated"] += 1
                else:
                    result["inserted"] += 1
            self._finish_batch(conn, batch_id, result)
        self.recompute_attributions()
        return result

    def replace_order_rows(self, rows, filename: str) -> dict[str, int]:
        with self.session() as conn:
            conn.execute("DELETE FROM order_attributions")
            conn.execute("DELETE FROM order_records")
        return self.import_order_rows(rows, filename)

    def set_manual_override(self, order_id: str, employee_user_id: str, owner_name_value: str, reason: str = "") -> None:
        with self.session() as conn:
            conn.execute(
                """
                INSERT INTO order_attribution_overrides(order_id, employee_user_id, owner_name, reason)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(order_id) DO UPDATE SET
                    employee_user_id = excluded.employee_user_id,
                    owner_name = excluded.owner_name,
                    reason = excluded.reason,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (order_id, employee_user_id, owner_name_value, reason),
            )
        self.recompute_attributions()

    def recompute_attributions(self) -> None:
        with self.session() as conn:
            orders = conn.execute("SELECT * FROM order_records").fetchall()
            for order in orders:
                override = conn.execute(
                    "SELECT * FROM order_attribution_overrides WHERE order_id = ?",
                    (order["order_id"],),
                ).fetchone()
                if override:
                    attribution = {
                        "union_id": order["wxid"],
                        "carry_record_id": None,
                        "employee_user_id": override["employee_user_id"],
                        "owner_name": override["owner_name"],
                        "channel": "",
                        "attribution_type": "manual",
                        "reason": override["reason"] or "手动改归因",
                    }
                else:
                    carry = conn.execute(
                        """
                        SELECT *
                        FROM carry_records
                        WHERE union_id = ?
                          AND add_time <= ?
                        ORDER BY add_time DESC, id DESC
                        LIMIT 1
                        """,
                        (order["wxid"], order["order_time"] or "9999-12-31 23:59:59"),
                    ).fetchone()
                    if carry:
                        channel = attribution_channel(carry["channel"], carry["customer_tags"])
                        attribution = {
                            "union_id": carry["union_id"],
                            "carry_record_id": carry["id"],
                            "employee_user_id": carry["employee_user_id"],
                            "owner_name": carry["owner_name"],
                            "channel": apply_manual_remark_channel(channel, order["manual_remark"]),
                            "attribution_type": "auto",
                            "reason": "最近一次添加自动归因",
                        }
                    elif is_activation_promoter(order["promoter_modified"] or order["promoter"]):
                        attribution = {
                            "union_id": order["wxid"],
                            "carry_record_id": None,
                            "employee_user_id": "",
                            "owner_name": "激活组",
                            "channel": "激活组",
                            "attribution_type": "activation_group",
                            "reason": "待补录订单推广员（修改后）=常丁健",
                        }
                    else:
                        attribution = {
                            "union_id": order["wxid"],
                            "carry_record_id": None,
                            "employee_user_id": "",
                            "owner_name": "待归因",
                            "channel": apply_manual_remark_channel("待归因", order["manual_remark"]),
                            "attribution_type": "pending",
                            "reason": "未找到订单时间之前的承接记录",
                        }
                conn.execute(
                    """
                    INSERT INTO order_attributions(
                        order_id, union_id, carry_record_id, employee_user_id,
                        owner_name, channel, attribution_type, reason
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(order_id) DO UPDATE SET
                        union_id = excluded.union_id,
                        carry_record_id = excluded.carry_record_id,
                        employee_user_id = excluded.employee_user_id,
                        owner_name = excluded.owner_name,
                        channel = excluded.channel,
                        attribution_type = excluded.attribution_type,
                        reason = excluded.reason,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (
                        order["order_id"],
                        attribution["union_id"],
                        attribution["carry_record_id"],
                        attribution["employee_user_id"],
                        attribution["owner_name"],
                        attribution["channel"],
                        attribution["attribution_type"],
                        attribution["reason"],
                    ),
                )

    def summary(self) -> dict[str, Any]:
        with self.session() as conn:
            carry_records = conn.execute("SELECT COUNT(*) AS count FROM carry_records").fetchone()["count"]
            unique_users = conn.execute("SELECT COUNT(DISTINCT union_id) AS count FROM carry_records").fetchone()["count"]
            orders = conn.execute("SELECT COUNT(*) AS count FROM order_records").fetchone()["count"]
            sales = conn.execute("SELECT COALESCE(SUM(net_sales), 0) AS total FROM order_records").fetchone()["total"]
            refunds = conn.execute("SELECT COALESCE(SUM(refund_amount), 0) AS total FROM order_records").fetchone()["total"]
            pending = conn.execute(
                "SELECT COUNT(*) AS count FROM order_attributions WHERE attribution_type = 'pending'"
            ).fetchone()["count"]
            manual = conn.execute(
                "SELECT COUNT(*) AS count FROM order_attributions WHERE attribution_type = 'manual'"
            ).fetchone()["count"]
            conversion_users = conn.execute(
                """
                SELECT COUNT(DISTINCT o.wxid) AS count
                FROM order_records o
                WHERE o.refund_amount < o.paid_amount
                """
            ).fetchone()["count"]
            return {
                "carry_records": int(carry_records),
                "unique_users": int(unique_users),
                "orders": int(orders),
                "sales": float(sales or 0),
                "refunds": float(refunds or 0),
                "pending_attributions": int(pending),
                "manual_attributions": int(manual),
                "conversion_users": int(conversion_users),
            }

    def carry_details(self, limit: int | None = None) -> list[dict[str, Any]]:
        sql = """
            SELECT
                id,
                union_id,
                employee_user_id,
                add_time,
                owner_raw,
                owner_name,
                channel,
                customer_tags
            FROM carry_records
            ORDER BY add_time DESC, id DESC
        """
        params: tuple[Any, ...] = ()
        if limit is not None:
            sql += " LIMIT ?"
            params = (limit,)
        with self.session() as conn:
            rows = conn.execute(sql, params).fetchall()
            return [dict(row) for row in rows]

    def owner_breakdown(self) -> list[dict[str, Any]]:
        with self.session() as conn:
            rows = conn.execute(
                """
                SELECT
                    COALESCE(a.owner_name, '待归因') AS owner_name,
                    COUNT(o.order_id) AS orders,
                    COALESCE(SUM(o.net_sales), 0) AS sales,
                    COUNT(DISTINCT CASE WHEN o.refund_amount < o.paid_amount THEN o.wxid END) AS conversion_users
                FROM order_records o
                LEFT JOIN order_attributions a ON a.order_id = o.order_id
                GROUP BY COALESCE(a.owner_name, '待归因')
                ORDER BY sales DESC
                """
            ).fetchall()
            return [dict(row) for row in rows]

    def batches(self) -> list[dict[str, Any]]:
        with self.session() as conn:
            rows = conn.execute(
                """
                SELECT *
                FROM import_batches
                ORDER BY id DESC
                LIMIT 50
                """
            ).fetchall()
            return [dict(row) for row in rows]

    def order_details(self, limit: int = 200) -> list[dict[str, Any]]:
        with self.session() as conn:
            rows = conn.execute(
                """
                SELECT
                    o.order_id,
                    o.order_no,
                    o.wxid,
                    o.user_id,
                    o.order_time,
                    o.paid_amount,
                    o.refund_amount,
                    o.net_sales,
                    o.order_status,
                    o.product_name,
                    o.scenario,
                    COALESCE(a.employee_user_id, '') AS employee_user_id,
                    COALESCE(a.owner_name, '待归因') AS owner_name,
                    COALESCE(a.channel, '待归因') AS channel,
                    COALESCE(a.attribution_type, 'pending') AS attribution_type,
                    COALESCE(a.reason, '') AS attribution_reason,
                    COALESCE(c.add_time, '') AS add_time,
                    COALESCE(c.owner_raw, '') AS owner_raw,
                    COALESCE(c.customer_tags, '') AS customer_tags
                FROM order_records o
                LEFT JOIN order_attributions a ON a.order_id = o.order_id
                LEFT JOIN carry_records c ON c.id = a.carry_record_id
                ORDER BY o.order_time DESC, o.order_id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
            return [dict(row) for row in rows]
